import streamlit as st
import psycopg2
import pandas as pd
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# code to run in databas+-

# CREATE TABLE chat_history (
#     id SERIAL PRIMARY KEY,
#     username VARCHAR(50),
#     question TEXT,
#     response TEXT,
#     timestamp TIMESTAMP
# );

# Function to connect to the PostgreSQL database
def connect_to_db():
    return psycopg2.connect(
        host=os.getenv("host"),
        database=os.getenv("database"),
        user=os.getenv("user"), 
        password=os.getenv("password"),
        port = os.getenv("port")
    )
# Fetch chats for specific user
def fetch_user_chats(username):
    conn = connect_to_db()
    cur = conn.cursor()
    cur.execute("SELECT question, response FROM chat_history WHERE username = %s ORDER BY timestamp DESC", (username,))
    chats = cur.fetchall()
    conn.close()
    return chats

# Store chat in DB
def store_chat(username, question, response):
    conn = connect_to_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO chat_history (username, question, response, timestamp) VALUES (%s, %s, %s, %s)",
                (username, question, response, datetime.now()))
    conn.commit()
    conn.close()

def sidebar():
    st.sidebar.title("DB Details")
    with st.sidebar.form(key='db_details'):
        st.text_input('Host', key="db_host")
        st.text_input('Username', key="db_user")
        st.text_input('Password', key="db_password", type='password')
        st.text_input('Port', key="db_port")
        st.text_input('Database', key="db_name")
        submit_button = st.form_submit_button(label='Submit')
        if submit_button:
            st.success("Database details updated!")
def fetch_data_from_postgres():
    conn = psycopg2.connect(
        host=os.getenv("host"),
        database=os.getenv("database"),
        user=os.getenv("user"), 
        password=os.getenv("password"),
        port = os.getenv("port")
    )
    query = "SELECT * FROM student"  # Adjust this to match your table
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# Function to convert the DataFrame to Excel in memory
# def convert_df_to_excel(df):
#     output = BytesIO()
#     writer = pd.ExcelWriter(output, engine='openpyxl')
#     df.to_excel(writer, index=False, sheet_name='Data')
#     writer.close()  # Use close() instead of save()
#     processed_data = output.getvalue()
#     return processed_data

def convert_df_to_excel(df):
    # Create a BytesIO stream to save the Excel file
    output = BytesIO()
    
    # Create a Pandas ExcelWriter object using openpyxl
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Write the DataFrame to the Excel file
    df.to_excel(writer, index=False, sheet_name='Data')
    
    # Access the workbook and the active worksheet
    workbook = writer.book
    worksheet = writer.sheets['Data']
    
    # Adjust column widths based on the maximum length of data in each column
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2  # Add padding
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Create a border style
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells in the data range
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Close the Pandas ExcelWriter and save to BytesIO
    writer.close()
    
    # Retrieve the Excel file content from the BytesIO stream
    processed_data = output.getvalue()  
    return processed_data

# Main function to handle user session and chat interface
def chat_interface():
    st.title("Chat Interface")
    with st.sidebar:
        sidebar()
        username = str(datetime.today().date())
        # username = st.text_input("Enter your username:", "User1")
        
        if st.button("View Previous Chats"):
            st.session_state.viewing_history = True

        if st.button("Chat with AI"):
            st.session_state.viewing_history = False

    # Check session state to toggle between views
    if "viewing_history" not in st.session_state:
        st.session_state.viewing_history = False

    # Toggle between viewing chat history and active chat interface
    if st.session_state.viewing_history:
        st.subheader("Chat History")
        chats = fetch_user_chats(username)
        if chats:
            for chat in chats:
                st.write(f"**You:** {chat[0]}")
                st.write(f"**Bot:** {chat[1]}")
        else:
            st.write("No previous chats available.")
    else:
        st.subheader("Chat with AI")
        # Chat input interface
        question = st.text_input("Ask a question:")
        if st.button("Send"):
            if question:
                response = "This is a placeholder response."  # Replace with actual bot logic
                df = fetch_data_from_postgres()
                if not df.empty:
                    st.write("Here is the data from the table:")
                    st.dataframe(df)

                    # Download as Excel button
                    st.download_button(
                        label="Download data as Excel",
                        data=convert_df_to_excel(df),
                        file_name="chat_data.xlsx",
                        mime="application/vnd.ms-excel"
                    )
                store_chat(username, question, response)
                st.write(f"**You:** {question}")
                st.write(f"**Bot:** {response}")
        
        # Show current chat history for this session
        chats = fetch_user_chats(username)  # Fetch chat history for current session
        if chats:
            for chat in chats:
                st.write(f"**You:** {chat[0]}")
                st.write(f"**Bot:** {chat[1]}")
                 

if __name__ == "__main__":
    load_dotenv()
    chat_interface()
