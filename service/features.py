from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd

class Features :

    @classmethod
    def convert_df_to_excel(cls,df):
        """
        Method to convert the dataframe into the excel file
        input : pandas DataFrame
        output : Bytes data
        """
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Data')
        
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2 
            worksheet.column_dimensions[column].width = adjusted_width
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        writer.close()
        
        processed_data = output.getvalue()  
        return processed_data
